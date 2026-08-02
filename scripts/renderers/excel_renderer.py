"""Excel 报告渲染器（openpyxl）。

生成多 Sheet Excel：
- 汇总：决策三态统计
- 决策明细：每个 ASIN 的决策结果
- 指标明细：五维得分
- 利润明细：利润分析
"""

from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# 决策状态颜色
DECISION_FILLS = {
    "GO": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "WATCH": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "NO-GO": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")


def _write_header(ws, headers: List[str]):
    """写表头并加样式。"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    # 自适应列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def _write_rows(ws, rows: List[List[Any]], start_row: int = 2):
    """写数据行。"""
    for row_offset, row_data in enumerate(rows):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=start_row + row_offset, column=col_idx, value=val)


def render_excel(
    rows: List[Dict[str, Any]],
    metrics_rows: List[Dict[str, Any]] = None,
    profit_rows: List[Dict[str, Any]] = None,
    category: str = "全部类目",
    cycle: str = "本周",
    output_path: str = None,
) -> str:
    """渲染 Excel 报告并写盘。

    Args:
        rows: 决策明细行
        metrics_rows: 指标明细行（可选）
        profit_rows: 利润明细行（可选）
        category: 类目名称
        cycle: 周期标签
        output_path: 输出文件路径

    Returns:
        输出文件路径
    """
    wb = Workbook()

    # Sheet 1: 汇总
    ws_summary = wb.active
    ws_summary.title = "汇总"
    total = len(rows)
    go_count = sum(1 for r in rows if r.get("decision") == "GO")
    watch_count = sum(1 for r in rows if r.get("decision") == "WATCH")
    nogo_count = sum(1 for r in rows if r.get("decision") == "NO-GO")
    _write_header(ws_summary, ["指标", "值"])
    _write_rows(ws_summary, [
        ["类目", category],
        ["周期", cycle],
        ["商品总数", total],
        ["GO 推荐", go_count],
        ["WATCH 观察", watch_count],
        ["NO-GO 不推荐", nogo_count],
        ["GO 占比", f"{(go_count/total*100):.1f}%" if total else "0.0%"],
    ])

    # Sheet 2: 决策明细
    ws_decision = wb.create_sheet("决策明细")
    decision_headers = ["ASIN", "决策", "综合得分", "风险得分", "利润率", "原因", "失败原因"]
    _write_header(ws_decision, decision_headers)
    decision_rows = []
    for r in rows:
        margin = r.get("profit_margin")
        decision_rows.append([
            r.get("asin", ""),
            r.get("decision", ""),
            r.get("composite_score", ""),
            r.get("risk_score", ""),
            f"{margin*100:.1f}%" if margin is not None else "—",
            r.get("reason", ""),
            r.get("failure_reason", ""),
        ])
    _write_rows(ws_decision, decision_rows)
    # 决策列染色
    for row_idx in range(2, len(decision_rows) + 2):
        decision = ws_decision.cell(row=row_idx, column=2).value
        if decision in DECISION_FILLS:
            ws_decision.cell(row=row_idx, column=2).fill = DECISION_FILLS[decision]

    # Sheet 3: 指标明细
    if metrics_rows:
        ws_metrics = wb.create_sheet("指标明细")
        headers = list(metrics_rows[0].keys())
        _write_header(ws_metrics, headers)
        _write_rows(ws_metrics, [
            [r.get(h, "") for h in headers] for r in metrics_rows
        ])

    # Sheet 4: 利润明细
    if profit_rows:
        ws_profit = wb.create_sheet("利润明细")
        headers = list(profit_rows[0].keys())
        _write_header(ws_profit, headers)
        _write_rows(ws_profit, [
            [r.get(h, "") for h in headers] for r in profit_rows
        ])

    if output_path:
        wb.save(output_path)

    return output_path
