"""报告生成器单元测试。"""

import os
import tempfile
import pytest
from report_generator import ReportGenerator
from renderers.html_renderer import render_html
from renderers.excel_renderer import render_excel


SAMPLE_ROWS = [
    {
        "asin": "B001",
        "decision": "GO",
        "composite_score": 85.0,
        "risk_score": 80.0,
        "profit_margin": 0.20,
        "reason": "全部达标",
        "failure_reason": "",
    },
    {
        "asin": "B002",
        "decision": "WATCH",
        "composite_score": 60.0,
        "risk_score": 70.0,
        "profit_margin": 0.10,
        "reason": "利润率 10.0% < 15.0%",
        "failure_reason": "",
    },
    {
        "asin": "B003",
        "decision": "NO-GO",
        "composite_score": 25.0,
        "risk_score": 20.0,
        "profit_margin": -0.05,
        "reason": "综合得分 25.0 < 40; 风险得分 20.0 < 30; 利润率 -5.0% < 0.0%",
        "failure_reason": "",
    },
    {
        "asin": "B004",
        "decision": "WATCH",
        "composite_score": 80.0,
        "risk_score": 80.0,
        "profit_margin": None,
        "reason": "利润率数据缺失",
        "failure_reason": "利润率数据缺失，无法判定 GO",
    },
]


class TestHtmlRenderer:
    def test_render_html_basic(self):
        html = render_html(SAMPLE_ROWS, category="Kitchen", cycle="本周")
        assert "选品决策报告" in html
        assert "B001" in html
        assert "GO" in html
        assert "WATCH" in html
        assert "NO-GO" in html

    def test_render_html_empty(self):
        html = render_html([], category="All", cycle="本周")
        assert "暂无决策数据" in html
        assert "商品总数" in html

    def test_render_html_writes_file(self):
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.html")
            render_html(SAMPLE_ROWS, output_path=path)
            assert os.path.exists(path)
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            assert "B001" in content

    def test_render_html_counts(self):
        html = render_html(SAMPLE_ROWS)
        # GO: 1, WATCH: 2, NO-GO: 1
        assert "1</div>" in html  # GO count
        # 检查占比显示
        assert "%" in html


class TestExcelRenderer:
    def test_render_excel_basic(self):
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.xlsx")
            render_excel(SAMPLE_ROWS, output_path=path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0

    def test_render_excel_with_metrics_and_profit(self):
        metrics = [{"ASIN": "B001", "综合得分": 85.0}]
        profit = [{"ASIN": "B001", "利润率": 0.20}]
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.xlsx")
            render_excel(
                SAMPLE_ROWS,
                metrics_rows=metrics,
                profit_rows=profit,
                output_path=path,
            )
            assert os.path.exists(path)

    def test_render_excel_empty(self):
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.xlsx")
            render_excel([], output_path=path)
            assert os.path.exists(path)

    def test_render_excel_sheets(self):
        """验证 Sheet 数量。"""
        from openpyxl import load_workbook
        metrics = [{"ASIN": "B001", "综合得分": 85.0}]
        profit = [{"ASIN": "B001", "利润率": 0.20}]
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.xlsx")
            render_excel(
                SAMPLE_ROWS,
                metrics_rows=metrics,
                profit_rows=profit,
                output_path=path,
            )
            wb = load_workbook(path)
            sheet_names = wb.sheetnames
            assert "汇总" in sheet_names
            assert "决策明细" in sheet_names
            assert "指标明细" in sheet_names
            assert "利润明细" in sheet_names


class TestReportGenerator:
    def test_run_generates_both_files(self):
        with tempfile.TemporaryDirectory() as d:
            gen = ReportGenerator(output_dir=d)
            result = gen.run(decision_rows=SAMPLE_ROWS, category="Kitchen", cycle="本周")
            assert "html" in result
            assert "excel" in result
            assert os.path.exists(result["html"])
            assert os.path.exists(result["excel"])

    def test_run_empty_rows(self):
        """空数据也能生成报告。"""
        with tempfile.TemporaryDirectory() as d:
            gen = ReportGenerator(output_dir=d)
            result = gen.run(decision_rows=[], category="All", cycle="本周")
            assert os.path.exists(result["html"])
            assert os.path.exists(result["excel"])

    def test_filename_safety(self):
        """类目名含特殊字符时文件名应安全。"""
        with tempfile.TemporaryDirectory() as d:
            gen = ReportGenerator(output_dir=d)
            result = gen.run(decision_rows=[], category="Kitchen & Dining/特价", cycle="本周")
            # 文件名应不含 / & 等特殊字符
            for p in result.values():
                assert "/" not in os.path.basename(p) or p.count("/") == 0
                assert "&" not in os.path.basename(p)
